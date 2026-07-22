"""
Rangam Infotech — Domestic Staffing BI Dashboard
ETL parser for the PERMANENT monthly template (confirmed unchanging):
  Sheets: Client List | Overall Revenue | Overall Performance |
          BOA | DB | MS | Baxter | Vantive | Grab | LSEG | WD+SD | Merck |
          TR Performance | ONB Performance

This replaces the earlier parser (which targeted the old single-sheet
"Clientwise" / "TR and ONB Performance" layout). Design principles:
  - APPEND-ONLY: this module never deletes anything. It returns which
    months are new vs. already in the database; the caller decides what
    to insert. A separate, explicit "replace month" flow (Data Management
    page) is the only path that overwrites existing data.
  - Same data-cleaning rules as before: mixed text/number revenue,
    blank-revenue = month not closed, '-' placeholders, free-text hours.
"""

import re
import datetime
import pandas as pd
import openpyxl

EXCEL_EPOCH = datetime.date(1899, 12, 30)

# Sheet name -> canonical client name / MSP name. Fixed mapping since the
# template is now confirmed permanent.
CLIENT_SHEET_MAP = {
    "BOA":    ("Bank of America", "Pontoon Solutions"),
    "DB":     ("Deutsche Bank", "Pontoon Solutions"),
    "MS":     ("Morgan Stanley", "Hays"),
    "Baxter": ("Baxter", "TAPFIN"),
    "Vantive":("Vantive", "Kelly OCG"),
    "Grab":   ("Grab H2O", "Pontoon Solutions"),
    "LSEG":   ("LSEG", "Hays"),
    "WD+SD":  ("Western Digital & Sandisk", "Magnit Global"),
    "Merck":  ("Merck", "Randstad"),
}


def excel_serial_to_date(v):
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        return EXCEL_EPOCH + datetime.timedelta(days=int(v))
    return None


def clean_currency(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = re.sub(r"[^\d.\-]", "", str(v))
    return round(float(s), 2) if s not in ("", "-", ".") else None


def parse_hours(v):
    if v is None:
        return None
    m = re.search(r"[\d.]+", str(v))
    return float(m.group()) if m else None


def parse_ratio(v):
    if v is None or v == "-":
        return None
    try:
        return round(float(v), 4)
    except (TypeError, ValueError):
        return None


def numval(v):
    return v if isinstance(v, (int, float)) else None


def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip().lower() if s is not None else ""


def fy_code(d):
    return "FY2026-27" if d >= datetime.date(2026, 4, 1) else "FY2025-26"


# ---------------------------------------------------------------------------
# Overall Performance -> company-wide monthly facts
# ---------------------------------------------------------------------------

def parse_overall_performance(wb):
    ws = wb["Overall Performance "]
    rows = []
    for r in range(3, ws.max_row + 1):
        month = excel_serial_to_date(ws.cell(r, 1).value)
        if month is None:
            # blank row or the 'Total' row — stop scanning further rows
            # only if we've already found at least one month, otherwise
            # keep looking in case of a stray blank row
            if rows:
                break
            continue
        rev_inr = clean_currency(ws.cell(r, 14).value)
        rev_usd = clean_currency(ws.cell(r, 15).value)
        rows.append({
            "month": month, "fy_code": fy_code(month),
            "new_reqs": ws.cell(r, 2).value, "worked_reqs": ws.cell(r, 3).value,
            "submissions": ws.cell(r, 4).value, "interviews": ws.cell(r, 5).value,
            "hire_preid": ws.cell(r, 6).value, "hire_sourced": ws.cell(r, 7).value,
            "start_preid": ws.cell(r, 8).value, "start_sourced": ws.cell(r, 9).value,
            "nothire_preid": ws.cell(r, 10).value, "nothire_sourced": ws.cell(r, 11).value,
            "concluded": ws.cell(r, 12).value, "headcount": ws.cell(r, 13).value,
            "revenue_inr": rev_inr, "revenue_usd": rev_usd,
            "is_month_closed": rev_inr is not None and rev_usd is not None,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Per-client sheets -> client-monthly facts
# ---------------------------------------------------------------------------

def parse_client_sheet(ws, client, msp):
    # Locate every column by its row-2 header text rather than a hardcoded
    # position. Merged category headers (Hire/Start/Not Hire) only populate
    # their first cell, so this naturally gives the start column of each
    # category. This also tolerates sheets missing columns entirely — e.g.
    # Vantive has no 'Hours Worked' / 'Avg. Recruiters' columns, which
    # shifts every column after them by 2 versus other client sheets.
    header_map = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(2, c).value)
        if h:
            header_map[h] = c

    def col(*name_options):
        for name in name_options:
            if name in header_map:
                return header_map[name]
        return None

    c_hours = col("hours worked")
    c_avgrec = col("avg. recruiters", "avg recruiters")
    c_newreqs = col("new reqs")
    c_workedreqs = col("worked reqs")
    c_subs = col("subs")
    c_ints = col("ints")
    c_hire = col("hire")
    c_start = col("start")
    c_nothire = col("not hire")
    c_concluded = col("concluded")
    c_headcount = col("headcount")
    c_revinr = col("revenue inr")
    c_revusd = col("revenue usd")

    def cell_val(r, c):
        return ws.cell(r, c).value if c else None

    # Split vs. combined Pre-Id/Sourced layout is detected by checking the
    # row directly below the 'Hire' header for a 'Pre-Id' sub-label.
    has_split = c_hire is not None and norm(ws.cell(3, c_hire).value) == "pre-id"
    data_start = 4 if has_split else 3

    rows = []
    for r in range(data_start, ws.max_row + 1):
        month = excel_serial_to_date(ws.cell(r, 1).value)
        if month is None:
            continue
        row = {
            "month": month, "fy_code": fy_code(month), "client": client, "msp": msp,
            "hours_worked": cell_val(r, c_hours), "avg_recruiters": cell_val(r, c_avgrec),
            "new_reqs": cell_val(r, c_newreqs), "worked_reqs": cell_val(r, c_workedreqs),
            "submissions": cell_val(r, c_subs), "interviews": cell_val(r, c_ints),
        }
        if has_split:
            row.update({
                "hire_preid": cell_val(r, c_hire), "hire_sourced": cell_val(r, c_hire + 1), "hire_combined": None,
                "start_preid": cell_val(r, c_start), "start_sourced": cell_val(r, c_start + 1), "start_combined": None,
                "nothire_preid": cell_val(r, c_nothire), "nothire_sourced": cell_val(r, c_nothire + 1), "nothire_combined": None,
            })
        else:
            row.update({
                "hire_preid": None, "hire_sourced": None, "hire_combined": cell_val(r, c_hire),
                "start_preid": None, "start_sourced": None, "start_combined": cell_val(r, c_start),
                "nothire_preid": None, "nothire_sourced": None, "nothire_combined": cell_val(r, c_nothire),
            })
        row["concluded"] = cell_val(r, c_concluded)
        row["headcount"] = cell_val(r, c_headcount)
        row["revenue_inr"] = clean_currency(cell_val(r, c_revinr))
        row["revenue_usd"] = clean_currency(cell_val(r, c_revusd))
        row["is_month_closed"] = row["revenue_inr"] is not None and row["revenue_usd"] is not None
        rows.append(row)
    return rows


def parse_all_clients(wb):
    all_rows = []
    for sheet_name, (client, msp) in CLIENT_SHEET_MAP.items():
        if sheet_name in wb.sheetnames:
            all_rows.extend(parse_client_sheet(wb[sheet_name], client, msp))
    return pd.DataFrame(all_rows)


# ---------------------------------------------------------------------------
# TR Performance / ONB Performance (separate sheets now)
# ---------------------------------------------------------------------------

def parse_tr_performance(wb, period_start, period_end):
    ws = wb["TR Performance"]
    # Column positions are located by header text (row 2) rather than
    # hardcoded indices — a 'Headcount' column has been seen inserted
    # before 'Revenue' in some files, which silently shifts everything
    # after it if positions are assumed fixed.
    header_map = {}
    for c in range(1, ws.max_column + 1):
        h = norm(ws.cell(2, c).value)
        if h:
            header_map[h] = c

    def col(*name_options, default=None):
        for name in name_options:
            if name in header_map:
                return header_map[name]
        return default

    c_new_reqs = col("new reqs", default=2)
    c_worked_reqs = col("worked reqs", default=3)
    c_subs = col("subs", default=4)
    c_ints = col("ints", default=5)
    c_hire = col("hire", default=6)
    c_start = col("start", default=7)
    c_not_hire = col("not hire", default=8)
    c_revenue = col("revenue", default=9)

    rows = []
    for r in range(3, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if label is None or norm(label) == "total":
            continue
        rows.append({
            "period_start": period_start, "period_end": period_end, "fy_code": fy_code(period_start),
            "recruiter_name": str(label).strip(),
            "is_pooled_bucket": norm(label) in ("pre-id", "left recruiters"),
            "new_reqs": numval(ws.cell(r, c_new_reqs).value), "worked_reqs": numval(ws.cell(r, c_worked_reqs).value),
            "submissions": numval(ws.cell(r, c_subs).value), "interviews": numval(ws.cell(r, c_ints).value),
            "hires": numval(ws.cell(r, c_hire).value), "starts": numval(ws.cell(r, c_start).value),
            "not_hires": numval(ws.cell(r, c_not_hire).value), "revenue": clean_currency(ws.cell(r, c_revenue).value),
        })
    return pd.DataFrame(rows)


def parse_onb_performance(wb, period_start, period_end):
    ws = wb["ONB Performance"]
    rows = []
    for r in range(4, ws.max_row + 1):  # row 3 is 'Individual Goal', skip
        label = ws.cell(r, 1).value
        if label is None or norm(label) == "total":
            continue
        rows.append({
            "period_start": period_start, "period_end": period_end, "fy_code": fy_code(period_start),
            "specialist_name": str(label).strip(),
            "hires": numval(ws.cell(r, 2).value), "starts": numval(ws.cell(r, 3).value),
            "avg_doc_completion_hours": parse_hours(ws.cell(r, 4).value),
            "avg_survey_completion_ratio": parse_ratio(ws.cell(r, 5).value),
            "avg_survey_satisfaction_ratio": parse_ratio(ws.cell(r, 6).value),
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Client List -> master data refresh
# ---------------------------------------------------------------------------

def parse_client_list(wb):
    ws = wb["Client List"]
    rows, current_msp = [], None
    for r in range(2, ws.max_row + 1):
        msp_cell = ws.cell(r, 1).value
        client_cell = ws.cell(r, 2).value
        if msp_cell:
            current_msp = str(msp_cell).strip()
        if client_cell:
            rows.append({
                "msp": current_msp, "client": str(client_cell).strip(),
                "start_date": excel_serial_to_date(ws.cell(r, 3).value),
                "duration_raw": ws.cell(r, 4).value,
            })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Validation — returns (errors, warnings). Errors BLOCK the import entirely;
# warnings are informational only.
# ---------------------------------------------------------------------------

REQUIRED_SHEETS = ["Client List", "Overall Revenue ", "Overall Performance ",
                   "TR Performance", "ONB Performance"] + list(CLIENT_SHEET_MAP.keys())


def validate_workbook(wb):
    errors = []
    warnings = []

    for s in REQUIRED_SHEETS:
        if s not in wb.sheetnames:
            errors.append(f"Missing required sheet: '{s}'")
    if errors:
        # Can't safely check header structure if sheets are missing entirely
        return errors, warnings

    # Overall Performance — check the header row has the expected columns
    # in the expected positions.
    ws = wb["Overall Performance "]
    expected_headers = {
        2: "new req", 3: "worked req", 4: "sub", 5: "int",
        14: "revenue", 15: "revenue",
    }
    for col, expect_substr in expected_headers.items():
        header_val = norm(ws.cell(1, col).value)
        if expect_substr not in header_val:
            errors.append(f"'Overall Performance' column {col} header is '{ws.cell(1, col).value}', "
                           f"expected something containing '{expect_substr}'. The column layout may have changed.")

    # At least one real monthly data row with a valid date
    has_date_row = any(excel_serial_to_date(ws.cell(r, 1).value) is not None for r in range(2, 10))
    if not has_date_row:
        errors.append("'Overall Performance' has no recognisable monthly date rows in the first few rows.")

    # Each client sheet — check it has a date column and a New Reqs-like column
    for sheet_name in CLIENT_SHEET_MAP:
        cws = wb[sheet_name]
        has_date = any(excel_serial_to_date(cws.cell(r, 1).value) is not None for r in range(1, 6))
        if not has_date:
            errors.append(f"'{sheet_name}' sheet has no recognisable monthly date rows — expected dates in column A.")

    # TR Performance — check it has a header row with 'Recruiter' and numeric columns
    tr = wb["TR Performance"]
    tr_header = norm(tr.cell(2, 1).value)
    if "recruiter" not in tr_header:
        warnings.append(f"'TR Performance' row 2, column A is '{tr.cell(2, 1).value}' — expected 'Recruiter'. "
                         f"Recruiter names will still be parsed, but double-check this sheet looks right.")

    return errors, warnings


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_workbook(path):
    """Returns a dict of clean dataframes ready for the caller to diff
    against the database and insert (append-only). If the workbook fails
    hard validation, 'errors' is non-empty and every other value is empty —
    callers must check 'errors' before using anything else in the dict."""
    wb = openpyxl.load_workbook(path, data_only=True)
    errors, warnings = validate_workbook(wb)

    if errors:
        return {
            "overall": pd.DataFrame(), "clientwise": pd.DataFrame(), "client_list": pd.DataFrame(),
            "recruiters": pd.DataFrame(), "onboarding": pd.DataFrame(),
            "months_detected": [], "errors": errors, "warnings": warnings,
        }

    overall = parse_overall_performance(wb)
    clientwise = parse_all_clients(wb)
    client_list = parse_client_list(wb)

    months = sorted(overall["month"].unique()) if len(overall) else []
    if months:
        period_start, period_end = months[0], months[-1]
        next_month = (period_end.replace(day=28) + datetime.timedelta(days=4)).replace(day=1)
        period_end = next_month - datetime.timedelta(days=1)
    else:
        period_start = period_end = None

    recruiters = parse_tr_performance(wb, period_start, period_end) if period_start else pd.DataFrame()
    onboarding = parse_onb_performance(wb, period_start, period_end) if period_start else pd.DataFrame()

    return {
        "overall": overall, "clientwise": clientwise, "client_list": client_list,
        "recruiters": recruiters, "onboarding": onboarding,
        "months_detected": [str(m) for m in months], "errors": errors, "warnings": warnings,
    }


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        "/mnt/user-data/uploads/Domestic_Staffing_Monthly_Review_June_2026_1.xlsx"
    data = parse_workbook(path)
    print("Errors:", data["errors"] or "none")
    print("Warnings:", data["warnings"] or "none")
    print("Months detected:", data["months_detected"])
    if not data["errors"]:
        print("\nOverall:\n", data["overall"].to_string(index=False))
        print(f"\nClientwise ({len(data['clientwise'])} rows):\n",
              data["clientwise"][["month","msp","client","new_reqs","hire_sourced","hire_combined","revenue_usd","is_month_closed"]].to_string(index=False))
        print(f"\nRecruiters ({len(data['recruiters'])} rows):\n", data["recruiters"].to_string(index=False))
        print(f"\nOnboarding ({len(data['onboarding'])} rows):\n", data["onboarding"].to_string(index=False))
